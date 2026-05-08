import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create Workbook
workbook = Workbook()
blueprint_sheet = workbook.active
blueprint_sheet.title = 'Asset Blueprint'

# Read CSV and write to sheet
csv_path = 'visual_assets_blueprint.csv'
with open(csv_path, mode='r', encoding='utf-8') as f:
    reader = csv.reader(f)
    for row in reader:
        blueprint_sheet.append(row)

# Formatting Blueprint Sheet
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Format Header
for cell in blueprint_sheet[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment
    cell.border = border

# Adjust column widths
column_widths = [15, 15, 15, 45, 35, 25, 35]
for i, width in enumerate(column_widths):
    blueprint_sheet.column_dimensions[get_column_letter(i+1)].width = width

# Format Body
for row in blueprint_sheet.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border

# Auto-filter
blueprint_sheet.auto_filter.ref = blueprint_sheet.dimensions

# 2. Summary Sheet
summary_sheet = workbook.create_sheet('Photography Guide')

guide_title_font = Font(bold=True, size=14)
guide_header_font = Font(bold=True, size=12)
normal_font = Font(size=11)

summary_sheet['A1'] = "Survival72 Visual Asset Strategy & Photography Guide"
summary_sheet['A1'].font = guide_title_font

guide_data = [
    ["", ""],
    ["1. Brand Style Reference", "Synthesis of Spyderco & Victorinox Aesthetics"],
    ["Victorinox Style (The 'Mechanical' Side)", "Use for Hero and PDP shots. Focus on 'Jewelry-grade' lighting, wheeling/rotating product showcases, and highlighting clean steel textures. Atmosphere: High-end, technical, precise."],
    ["Spyderco Style (The 'Functional' Side)", "Use for Lifestyle and Action shots. Focus on 'Real-world' EDC environments (camping, engine repair), natural lighting, and showing the tools in actual hand use. Atmosphere: Rugged, reliable, authentic."],
    ["", ""],
    ["2. Visual Content Hierarchy", ""],
    ["Precision (Hero/System)", "Clean contour lighting, 45-degree angles, exploded views of modules."],
    ["Reliability (Problem/Macro)", "High-contrast close-ups, showing stress tests, steel grain, and mechanical 'bite' points."],
    ["Lifestyle (PDP/About)", "Authentic outdoors/workshop settings, warm tones, human connection (hand grip)."],
    ["", ""],
    ["3. Essential Visual Elements", ""],
    ["Color Accent", "Survival72 Orange (indicators/buttons) must pop against the cold steel backgrounds."],
    ["Material Texture", "High-def capture of steel brushing and CNC machining marks to justify premium pricing."],
    ["Branding", "Ensure Logo and Steel Grade markings are in sharp focus in every detail shot."],
]

for i, row_data in enumerate(guide_data, start=2):
    for j, val in enumerate(row_data, start=1):
        cell = summary_sheet.cell(row=i, column=j, value=val)
        if j == 1 and val != "":
            cell.font = guide_header_font
        else:
            cell.font = normal_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

summary_sheet.column_dimensions['A'].width = 30
summary_sheet.column_dimensions['B'].width = 80

# Final Save
excel_path = 'Survival72_Visual_Asset_Blueprint.xlsx'
workbook.save(excel_path)
print(f"Excel file created successfully: {excel_path}")
